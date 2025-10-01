"""Spreadsheet document extractor for CSV and Excel files."""

from __future__ import annotations

import csv
import logging
import time
from pathlib import Path
from typing import Any, ClassVar

import openpyxl
from files.backend.extractors.base import DocumentExtractor, ExtractionResult

logger = logging.getLogger(__name__)
MAX_TEXT_ROWS = 100


class SpreadsheetExtractor(DocumentExtractor):
    """Extract content from spreadsheet documents"""

    SUPPORTED_EXTENSIONS: ClassVar[set[str]] = {".xlsx", ".xls", ".csv", ".tsv"}

    async def can_extract(self, file_path: Path) -> bool:
        """Check if this extractor can handle the given file"""
        return file_path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    async def extract(self, file_path: Path, options: dict[str, Any] | None = None) -> ExtractionResult:
        """Extract content from spreadsheet"""
        start_time = time.time()
        options = options or {}

        # Validate file
        await self.validate_file(file_path)

        file_ext = file_path.suffix.lower()

        result = ExtractionResult(
            file_path=str(file_path),
            file_type=file_ext[1:],  # Remove the dot
            file_size=file_path.stat().st_size,
            extraction_method="pandas" if file_ext in [".xlsx", ".xls"] else "csv",
            processing_time_ms=0,
        )

        try:
            if file_ext in [".csv", ".tsv"]:
                tables = await self._extract_csv(file_path, file_ext == ".tsv")
            else:
                tables = await self._extract_excel(file_path)

            result.tables = tables

            # Generate title from filename
            result.title = file_path.stem

            # Create sections from tables
            sections = []
            for idx, table in enumerate(tables):
                section = {
                    "level": 1,
                    "title": table.get("name", f"Sheet {idx + 1}"),
                    "content": self._table_to_text(table),
                    "order": idx + 1,
                    "path": str(idx + 1),
                }
                sections.append(section)

            result.sections = sections

            # Generate full text
            result.full_text = "\n\n".join(s["content"] for s in sections)

        except Exception as e:
            logger.exception(f"Error extracting spreadsheet {file_path}")
            result.error_messages.append(str(e))

        result.processing_time_ms = int((time.time() - start_time) * 1000)
        return result

    async def _extract_csv(self, file_path: Path, is_tsv: bool = False) -> list[dict[str, Any]]:
        """Extract data from CSV/TSV file"""
        tables: list[dict[str, Any]] = []

        try:
            delimiter = "\t" if is_tsv else ","

            with file_path.open(encoding="utf-8-sig") as file:
                # Try to detect dialect
                sample = file.read(1024)
                file.seek(0)

                try:
                    dialect = csv.Sniffer().sniff(sample)
                    delimiter = dialect.delimiter
                except Exception:
                    logger.debug("Failed to detect CSV dialect, using default delimiter")

                reader = csv.DictReader(file, delimiter=delimiter)

                headers = list(reader.fieldnames) if reader.fieldnames else []
                if not headers:
                    return tables

                rows = list(reader)

                if rows:
                    table = {
                        "name": file_path.stem,
                        "headers": headers,
                        "rows": rows,
                        "schema": self.infer_table_schema(headers, rows),
                    }
                    tables.append(table)

        except Exception as exc:
            logger.error(f"Failed to extract CSV/TSV with csv module: {exc}")
            raise

        return tables

    async def _extract_excel(self, file_path: Path) -> list[dict[str, Any]]:
        """Extract data from Excel file"""
        tables: list[dict[str, Any]] = []

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise RuntimeError(f"Failed to open Excel workbook: {exc}") from exc

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                table = self._extract_openpyxl_sheet(sheet_name, sheet)
                if table:
                    tables.append(table)
        finally:
            workbook.close()

        return tables

    def _extract_openpyxl_sheet(self, sheet_name: str, sheet: Any) -> dict[str, Any] | None:
        """Extract table data from an openpyxl worksheet."""
        max_row = sheet.max_row
        max_col = sheet.max_column
        if max_row == 0 or max_col == 0:
            return None

        headers = [self._cell_to_header(sheet.cell(1, col).value, col) for col in range(1, max_col + 1)]
        rows = self._collect_sheet_rows(sheet, headers, max_row, max_col)
        if not rows:
            return None

        return {
            "name": sheet_name,
            "headers": headers,
            "rows": rows,
            "schema": self.infer_table_schema(headers, rows),
        }

    def _cell_to_header(self, value: Any, column_index: int) -> str:
        """Convert a header cell to string value."""
        return str(value) if value else f"Column_{column_index}"

    def _collect_sheet_rows(
        self,
        sheet: Any,
        headers: list[str],
        max_row: int,
        max_col: int,
    ) -> list[dict[str, Any]]:
        """Collect non-empty rows from an openpyxl worksheet."""
        rows: list[dict[str, Any]] = []
        for row_index in range(2, max_row + 1):
            row_data: dict[str, Any] = {}
            has_data = False
            for col_index in range(1, max_col + 1):
                cell_value = sheet.cell(row_index, col_index).value
                if cell_value is None:
                    continue
                has_data = True
                row_data[headers[col_index - 1]] = cell_value
            if has_data:
                rows.append(row_data)
        return rows

    def _table_to_text(self, table: dict[str, Any]) -> str:
        """Convert table to text representation"""
        lines = []

        # Add table name
        lines.append(f"Table: {table.get('name', 'Unnamed')}")
        lines.append("")

        # Add headers
        headers = table.get("headers", [])
        if headers:
            lines.append(" | ".join(headers))
            lines.append("-" * (len(" | ".join(headers))))

        # Add rows
        rows = table.get("rows", [])
        for row in rows[:MAX_TEXT_ROWS]:
            row_values = []
            for header in headers:
                value = row.get(header, "")
                row_values.append(str(value) if value is not None else "")
            lines.append(" | ".join(row_values))

        if len(rows) > MAX_TEXT_ROWS:
            lines.append(f"... and {len(rows) - MAX_TEXT_ROWS} more rows")

        return "\n".join(lines)
